#!/usr/bin/env python3
"""주간보고 엑셀 파일에서 한 주차 시트를 찾아 data/weekly_activity_<이름>.csv에 반영한다.

엑셀은 "OO월 N주" 형태의 시트가 주차별로 쌓여 있고, 각 시트는
이슈사항 / 전주 진척사항 / 금주 진척사항 / 비고 4개 컬럼의 표로 구성된다.

사용법:
    python scripts/extract_weekly_activity.py <이름> <엑셀파일경로> <weekOf>

    weekOf 예)
      - "8월2주", "26년8월2주" 처럼 시트명/제목에 포함된 주차 표기와
        매칭되는 문자열
      - "latest" : 엑셀에서 가장 마지막(가장 최근) 시트를 사용

    예) python scripts/extract_weekly_activity.py 심유선 ~/Downloads/주간보고.xlsx latest

동일한 weekOf(정규화된 주차 라벨 기준)로 재실행하면 기존 데이터를 새 결과로 교체한다(중복 누적 방지).
"""

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent

CSV_FIELDS = [
    "weekOf", "topic", "prev_week_progress", "this_week_progress",
    "status", "week_source", "source_file", "extracted_at",
]

HEADER_KEYWORDS = {
    "topic": ["이슈사항"],
    "prev": ["전주"],
    "this_week": ["금주"],
    "status": ["비고"],
}


def normalize(text) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", "", str(text))


def cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def find_title(rows):
    for row in rows:
        for cell in row:
            text = cell_text(cell)
            if "주간보고" in text:
                return text
    return ""


def week_label(sheet_name: str, title_text: str) -> str:
    """시트명/제목에서 연-월-주차를 뽑아 정렬 가능한 라벨을 만든다. 못 찾으면 시트명을 그대로 쓴다."""
    year_match = re.search(r"(\d{2,4})년", sheet_name) or re.search(r"(\d{2,4})년", title_text)
    mw_match = re.search(r"(\d{1,2})월\s*(\d{1,2})주", title_text) or re.search(r"(\d{1,2})월\s*(\d{1,2})주", sheet_name)
    if year_match and mw_match:
        year = year_match.group(1)
        year = "20" + year if len(year) == 2 else year
        month, week = mw_match.groups()
        return f"{year}-{int(month):02d}-W{int(week)}"
    return sheet_name.strip()


def find_header(rows):
    for r_idx, row in enumerate(rows):
        cols = {}
        for c_idx, cell in enumerate(row):
            text = cell_text(cell)
            if not text:
                continue
            if text == "이슈사항":
                cols["topic"] = c_idx
            elif "전주" in text:
                cols["prev"] = c_idx
            elif "금주" in text:
                cols["this_week"] = c_idx
            elif text == "비고":
                cols["status"] = c_idx
        if "topic" in cols and "this_week" in cols:
            return r_idx, cols
    return None, None


def extract_sheet_records(ws, weekOf_label: str, source_file: str):
    rows = list(ws.iter_rows(values_only=True))
    header_idx, cols = find_header(rows)
    if header_idx is None:
        return []

    records = []
    extracted_at = datetime.now().isoformat(timespec="seconds")
    for row in rows[header_idx + 1:]:
        topic = cell_text(row[cols["topic"]]) if cols["topic"] < len(row) else ""
        if not topic:
            continue
        prev = cell_text(row[cols["prev"]]) if "prev" in cols and cols["prev"] < len(row) else ""
        this_week = cell_text(row[cols["this_week"]]) if cols["this_week"] < len(row) else ""
        status = cell_text(row[cols["status"]]) if "status" in cols and cols["status"] < len(row) else ""

        if not (prev or this_week or status):
            continue

        records.append({
            "weekOf": weekOf_label,
            "topic": topic,
            "prev_week_progress": prev,
            "this_week_progress": this_week,
            "status": status,
            "week_source": ws.title.strip(),
            "source_file": source_file,
            "extracted_at": extracted_at,
        })
    return records


def resolve_sheet(wb, weekOf: str):
    if normalize(weekOf) in ("latest", "최신"):
        return wb.worksheets[-1]

    target = normalize(weekOf)
    candidates = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        title_text = find_title(rows)
        label = week_label(ws.title, title_text)
        keys = {normalize(ws.title), normalize(title_text), normalize(label)}
        if any(target == k or (target and target in k) for k in keys if k):
            candidates.append(ws)

    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        # 더 구체적으로(제목 텍스트까지) 일치하는 시트를 우선한다
        return candidates[-1]
    return None


def load_existing(csv_path: Path):
    if not csv_path.exists():
        return []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(csv_path: Path, rows):
    rows = sorted(rows, key=lambda r: (r.get("weekOf", ""), r.get("topic", "")))
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in CSV_FIELDS})


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("name", help="사육자/담당자 이름 (예: 심유선)")
    parser.add_argument("filepath", help="업로드된 주간보고 엑셀 파일 경로")
    parser.add_argument("weekOf", help="대상 주차 (예: 8월2주, 26년8월2주, latest)")
    args = parser.parse_args()

    filepath = Path(args.filepath).expanduser()
    if not filepath.exists():
        sys.exit(f"엑셀 파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = resolve_sheet(wb, args.weekOf)
    if ws is None:
        available = ", ".join(s.title.strip() for s in wb.worksheets)
        sys.exit(
            f"'{args.weekOf}'와(과) 일치하는 주차 시트를 찾지 못했습니다.\n"
            f"사용 가능한 시트: {available}\n"
            f"(또는 weekOf에 'latest'를 넣으면 가장 최근 시트를 사용합니다.)"
        )

    rows = list(ws.iter_rows(values_only=True))
    title_text = find_title(rows)
    weekOf_label = week_label(ws.title, title_text)

    new_records = extract_sheet_records(ws, weekOf_label, filepath.name)
    if not new_records:
        sys.exit(
            f"'{ws.title.strip()}' 시트에서 추출된 항목이 없습니다. "
            "'이슈사항/전주 진척사항/금주 진척사항/비고' 헤더가 있는지 확인해주세요."
        )

    csv_path = REPO_ROOT / "data" / f"weekly_activity_{args.name}.csv"
    existing = load_existing(csv_path)
    kept = [r for r in existing if r.get("weekOf") != weekOf_label]
    write_csv(csv_path, kept + new_records)

    print(f"'{ws.title.strip()}' 시트({weekOf_label}) 반영 완료 -> {csv_path.relative_to(REPO_ROOT)}")
    print(f"  {len(new_records)}개 이슈 항목 기록")


if __name__ == "__main__":
    main()
